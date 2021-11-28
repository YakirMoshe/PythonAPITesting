import json
import requests
import jsonpath
import openpyxl

def test_add_multiple_student():
    # API
    API_URL = "https://reqres.in/api/users"
    file = open('C:/TASK_API/AddNewStudent.json')
    # Excel Code
    vk = openpyxl.load_workbook('C:/TASK_API/TestData.xlsx')
    sh =vk['Sheet1']
    rows = sh.max_row
    json_request = json.loads(file.read())
    for i in range(2, rows+1):
         cell_name = sh.cell(row=i, column=1)
         cell_job = sh.cell(row=i, column=2)
         json_request['name'] = cell_name.value
         json_request['job'] = cell_job.value
         response = requests.post(API_URL, json_request)
         print(response.text)
         print(response.status_code)
         assert response.status_code == 200